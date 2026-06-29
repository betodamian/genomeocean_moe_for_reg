#!/usr/bin/env python3
"""
Parse all auto-fetched RBS and Rho label xlsx tables into standardised TSVs.

RBS output schema (per source):
  source_id, organism, strain, gene_name, locus_tag, tis_pos, tis_end,
  strand, start_codon, ribo_density, sd_motif, sd_pos_rel, evidence_type, notes

Rho output schema (per source):
  source_id, organism, strain, site_id, site_start, site_end, strand,
  site_class, terminates_or_silences, sequence, evidence_type, notes

Run: .venv/bin/python pipeline/parse_label_tables.py
"""
import csv, os, re
import openpyxl

ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RBS   = os.path.join(ROOT, "data/rbs_database/raw")
RHO   = os.path.join(ROOT, "data/rho_database/raw")
GEN   = os.path.join(ROOT, "data/genomes")

RBS_FIELDS = ["source_id","organism","strain","gene_name","locus_tag",
              "tis_pos","tis_end","strand","start_codon",
              "ribo_density","sd_motif","sd_pos_rel","evidence_type","notes"]
RHO_FIELDS = ["source_id","organism","strain","site_id",
              "site_start","site_end","strand","site_class",
              "terminates_or_silences","sequence","evidence_type","notes"]

# ── Controlled evidence vocabulary (per research_plan §12 tier-gating) ────────
# Headline ground truth = T1_* (in vivo). T2_* (in vitro) flagged separately.
# ALGO_* never used as headline ground truth (anti-circularity, §12).
# Applied centrally in write_tsv() so every source emits one normalized tier;
# method detail (BCM vs depletion, retapamulin vs tetracycline) stays in notes.
EVIDENCE_TIER = {
    # RBS — ribosome-profiling translation-initiation sites (T1 in vivo)
    "ECOLI_RIBORET":      "T1_riboseq",
    "ECOLI_BL21_RIBORET": "T1_riboseq",
    "ECOLI_TETRP":   "T1_riboseq",
    "MTB_RIBOSEQ":   "T1_riboseq",
    "BSUB_RIBOSEQ":  "T1_riboseq",
    "BSUB_SPORE":    "T1_riboseq",
    "CAULO_RIBOSEQ": "T1_riboseq",
    "SAUR_EXSD":     "T1_riboseq",
    "HVOLC_RIBOSEQ": "T1_riboseq",
    # Rho-dependent terminators — in vivo
    "ECOLI_BCM_RHO": "T1_invivo",
    "MTB_RHODUC":    "T1_invivo",
    "MTB_RHODUC2":   "T1_invivo",
    # Intrinsic terminators (experimentally confirmed; Rho decoy/contrast class)
    "INTRINSIC_TERMITE": "T1_termseq",
    # in vitro
    "BSUB_HSELEX":   "T2_invitro",
    # algorithm-predicted — cross-check only, never headline (§12)
    "RHOTERMPREDICT": "ALGO_xcheck",
}

def wb(path):
    return openpyxl.load_workbook(path, read_only=True)

def sheet_rows(workbook, sheet_name, skip=1):
    ws = workbook[sheet_name]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        if not any(c is not None for c in r):
            continue
        yield list(r)

def write_tsv(rows, fields, path):
    # Normalize evidence_type from the controlled vocabulary (single source of
    # truth) so the §12 tier-gate can rely on T1_/T2_/ALGO_ prefixes.
    if "evidence_type" in fields:
        for r in rows:
            tier = EVIDENCE_TIER.get(r.get("source_id"))
            if tier:
                r["evidence_type"] = tier
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, delimiter="\t", lineterminator="\n",
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {len(rows):>5,} rows -> {os.path.relpath(path)}")
    return len(rows)


# ── GFF helper ──────────────────────────────────────────────────────────────

def load_gff_locus_map(org_dir):
    """Return dict locus_tag -> (start_1based, end_1based, strand) from GFF CDS features."""
    gff = next(os.path.join(org_dir, f)
               for f in os.listdir(org_dir) if f.endswith(".gff"))
    mapping = {}
    with open(gff) as fh:
        for ln in fh:
            if ln.startswith("#") or "\t" not in ln:
                continue
            cols = ln.rstrip("\n").split("\t")
            if len(cols) < 9 or cols[2] != "CDS":
                continue
            attrs = cols[8]
            m = re.search(r'locus_tag=([^;]+)', attrs)
            if not m:
                continue
            tag = m.group(1)
            mapping[tag] = (int(cols[3]), int(cols[4]), cols[6])
    return mapping


# ════════════════════════════════════════════════════════════════════════════
# RBS PARSERS
# ════════════════════════════════════════════════════════════════════════════

def parse_ecoli_tetrp():
    """Nakahigashi 2016 TetRP: novel TIS (Table S2) + reannotated starts (Table S3)."""
    src  = "ECOLI_TETRP"
    path = os.path.join(RBS, src, "supp_dsw008_dsw008supp_table2-5.xlsx")
    rows = []
    book = wb(path)

    # Table_S2: novel TIS calls
    # cols: No, ECK, gene_name, dir, new_initiation, shift, new_start_codon, ...
    for r in sheet_rows(book, "Table_S2", skip=1):
        if r[0] is None:
            continue
        gene    = str(r[2]).strip() if r[2] else None
        strand  = str(r[3]).strip() if r[3] else None
        tis_pos = int(r[4]) if r[4] is not None else None
        codon   = str(r[6]).strip() if r[6] else None
        locus   = str(r[1]).strip() if r[1] else None
        if tis_pos is None:
            continue
        rows.append(dict(
            source_id="ECOLI_TETRP", organism="ecoli_K12_MG1655",
            strain="BW25113", gene_name=gene, locus_tag=locus,
            tis_pos=tis_pos, tis_end=None, strand=strand,
            start_codon=codon, ribo_density=None,
            sd_motif=None, sd_pos_rel=None,
            evidence_type="T1_tetrp",
            notes=f"novel_TIS;shift_from_annotated={r[5]}",
        ))

    # Table_S3: reannotated gene starts
    # cols: EcoGene_ID, ECK_ID, Figure, Gene_name, Dir, Left(2014), Right(2014), Left(corrected)
    for r in sheet_rows(book, "Table_S3", skip=1):
        if r[0] is None:
            continue
        gene    = str(r[3]).strip() if r[3] else None
        strand  = str(r[4]).strip() if r[4] else None
        tis_pos = int(r[7]) if r[7] is not None else None
        locus   = str(r[1]).strip() if r[1] else None
        if tis_pos is None:
            continue
        rows.append(dict(
            source_id="ECOLI_TETRP", organism="ecoli_K12_MG1655",
            strain="BW25113", gene_name=gene, locus_tag=locus,
            tis_pos=tis_pos, tis_end=None, strand=strand,
            start_codon=None, ribo_density=None,
            sd_motif=None, sd_pos_rel=None,
            evidence_type="T1_tetrp",
            notes="reannotated_start",
        ))

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


def parse_caulo_riboseq():
    """Schrader 2014: C. crescentus NA1000 annotated CDS positions from ribosome profiling."""
    src  = "CAULO_RIBOSEQ"
    path = os.path.join(RBS, src, "pgen.1004463.s018.xlsx")
    rows = []
    book = wb(path)

    # CDSs sheet: ORF_ID, Locus_ID, Left_coord, Right_coord, Strand, Gene_ID, Gene_product
    for r in sheet_rows(book, "CDSs", skip=1):
        if r[0] is None:
            continue
        locus  = str(r[1]).strip() if r[1] else None
        left   = int(r[2]) if r[2] is not None else None
        right  = int(r[3]) if r[3] is not None else None
        strand = str(r[4]).strip() if r[4] else None
        gene   = str(r[5]).strip() if r[5] else None
        if left is None or strand is None:
            continue
        tis_pos = left  if strand == "+" else right
        tis_end = right if strand == "+" else left
        rows.append(dict(
            source_id="CAULO_RIBOSEQ", organism="ccrescentus_NA1000",
            strain="NA1000", gene_name=gene, locus_tag=locus,
            tis_pos=tis_pos, tis_end=tis_end, strand=strand,
            start_codon=None, ribo_density=None,
            sd_motif=None, sd_pos_rel=None,
            evidence_type="T1_riboseq",
            notes="annotated_CDS",
        ))

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


def parse_saur_exsd():
    """Kohl 2026: S. aureus HG001 novel sORFs with extended-SD annotations."""
    src  = "SAUR_EXSD"
    path = os.path.join(RBS, src, "41467_2026_69079_MOESM3_ESM.xlsx")
    rows = []
    book = wb(path)

    # Sheet1: HG001_sORF_ID, Start, End, Strand, SD_prediction, Length_aa, Sequence, Start_codon
    # Row 0 = title, row 1 = header
    for r in sheet_rows(book, "Sheet1", skip=2):
        if r[0] is None:
            continue
        start  = int(r[1]) if r[1] is not None else None
        end    = int(r[2]) if r[2] is not None else None
        strand = str(r[3]).strip() if r[3] else None
        sd_raw = str(r[4]).strip() if r[4] else None
        codon  = str(r[7]).replace("U", "T").strip() if r[7] else None
        if start is None or strand is None:
            continue
        tis_pos = start if strand == "+" else end
        tis_end = end   if strand == "+" else start

        # Parse SD motif and position from "AGGA_-12" or "GGGG_-11;AGGA_-4;..."
        sd_motif = sd_pos = None
        if sd_raw:
            m = re.match(r'([A-Z]+)_([-\d]+)', sd_raw)
            if m:
                sd_motif = m.group(1)
                sd_pos   = int(m.group(2))

        rows.append(dict(
            source_id="SAUR_EXSD", organism="saureus_HG001",
            strain="HG001", gene_name=None, locus_tag=f"HG001_sORF_{r[0]}",
            tis_pos=tis_pos, tis_end=tis_end, strand=strand,
            start_codon=codon, ribo_density=None,
            sd_motif=sd_motif, sd_pos_rel=sd_pos,
            evidence_type="T1_riboseq",
            notes=f"novel_sORF;sd_raw={sd_raw};coords=HG001_CP018205.1;"
                  f"also_in_SAUR_RIBORET_peakcall",
        ))

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


def parse_mtb_riboseq():
    """Sawyer 2021: MTB H37Rv ribosome occupancy per gene; cross-ref GFF for coordinates."""
    src  = "MTB_RIBOSEQ"
    path = os.path.join(RBS, src, "mmc2.xlsx")
    rows = []
    book = wb(path)

    # Build locus_tag -> (start, end, strand) from MTB GFF
    gff_map = load_gff_locus_map(os.path.join(GEN, "mtuberculosis_H37Rv"))

    # Table S1C: gene(=locus_tag), counts rep1-3, sum, RO rep1-3
    # Skip first 3 rows (title/blank/header)
    for r in sheet_rows(book, "Table S1C", skip=3):
        if not r[0]:
            continue
        tag = str(r[0]).strip()
        # Parse ribosome occupancy (mean of 3 reps, skip NaN/formula cells)
        ro_vals = []
        for col in [5, 6, 7]:
            try:
                v = float(r[col]) if r[col] is not None else None
                if v is not None:
                    ro_vals.append(v)
            except (ValueError, TypeError):
                pass
        ribo = sum(ro_vals) / len(ro_vals) if ro_vals else None

        coords = gff_map.get(tag)
        if coords is None:
            # Try Rv tag variants (e.g., trailing 'c')
            coords = gff_map.get(tag.rstrip("c")) or gff_map.get(tag + "c")
        if coords:
            start, end, strand = coords
            tis_pos = start if strand == "+" else end
            tis_end = end   if strand == "+" else start
        else:
            tis_pos = tis_end = strand = None

        rows.append(dict(
            source_id="MTB_RIBOSEQ", organism="mtuberculosis_H37Rv",
            strain="H37Rv", gene_name=None, locus_tag=tag,
            tis_pos=tis_pos, tis_end=tis_end, strand=strand,
            start_codon=None, ribo_density=ribo,
            sd_motif=None, sd_pos_rel=None,
            evidence_type="T1_riboseq",
            notes="gff_coords" if tis_pos else "no_gff_match",
        ))

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


def parse_bsub_spore():
    """Iwańska 2024: B. subtilis sporulation ribosome profiling; use T0 (vegetative) sheet."""
    src  = "BSUB_SPORE"
    path = os.path.join(RBS, src, "41467_2024_51654_MOESM7_ESM.xlsx")
    rows = []
    book = wb(path)

    # T0 (vegetative) gives the broadest gene set and most comparable to BSUB_RIBOSEQ
    # Columns: BSGatlas_ID, locus(BSU_xxxx), name, start, end, strand, WT0r_1, WT0r_2, WT0r_3...
    for r in sheet_rows(book, "T0", skip=1):
        if not r[0]:
            continue
        locus  = str(r[1]).strip() if r[1] else None
        name   = str(r[2]).strip() if r[2] else None
        start  = int(r[3]) if r[3] is not None else None
        end    = int(r[4]) if r[4] is not None else None
        strand = str(r[5]).strip() if r[5] else None
        # WT ribosome counts: cols 6, 7, 8 (WT0r_1, WT0r_2, WT0r_3 if present)
        counts = []
        for col in range(6, min(9, len(r))):
            try:
                v = float(r[col]) if r[col] is not None else None
                if v is not None:
                    counts.append(v)
            except (ValueError, TypeError):
                pass
        density = sum(counts) / len(counts) if counts else None

        if start is None or strand is None:
            continue
        tis_pos = start if strand == "+" else end
        tis_end = end   if strand == "+" else start

        rows.append(dict(
            source_id="BSUB_SPORE", organism="bsubtilis_168",
            strain="168", gene_name=name, locus_tag=locus,
            tis_pos=tis_pos, tis_end=tis_end, strand=strand,
            start_codon=None, ribo_density=density,
            sd_motif=None, sd_pos_rel=None,
            evidence_type="T1_riboseq",
            notes="T0_vegetative_growth",
        ))

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# RHO PARSERS
# ════════════════════════════════════════════════════════════════════════════

def parse_mtb_rhoduc():
    """Botella 2017: MTB Rho-sensitive regions (RSRs) with class and termination type."""
    src  = "MTB_RHODUC"
    path = os.path.join(RHO, src, "ncomms14731-s3.xlsx")
    rows = []
    book = wb(path)

    # Sheet '2-RSRs': RSR_ID, RSR_Start, RSR_End, RSR_length, Class, Term(T)/Silence(S), ...
    # Skip 4 rows (title, note, blank, header)
    for r in sheet_rows(book, "2-RSRs", skip=4):
        if r[0] is None:
            continue
        rsr_id  = r[0]
        s1, s2  = r[1], r[2]
        rclass  = str(r[4]).strip() if r[4] else None
        ts      = str(r[5]).strip() if r[5] else None
        if s1 is None or s2 is None:
            continue
        # Strand: positive RSR_ID = + strand, negative = - strand
        strand  = "+" if (isinstance(rsr_id, (int, float)) and rsr_id > 0) else "-"
        # Normalise coordinates (minus-strand entries can be reversed)
        start   = int(min(s1, s2))
        end     = int(max(s1, s2))

        rows.append(dict(
            source_id="MTB_RHODUC", organism="mtuberculosis_H37Rv",
            strain="H37Rv", site_id=f"RSR_{rsr_id}",
            site_start=start, site_end=end, strand=strand,
            site_class=rclass, terminates_or_silences=ts,
            sequence=None,
            evidence_type="T1_invivo_genetic_depletion",
            notes="RhoDUC_Rho_sensitive_region",
        ))

    book.close()
    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


def parse_bsub_hselex():
    """H-SELEX rut peaks for B. subtilis (BsRho) and E. coli (EcRho) — T2 in vitro."""
    src  = "BSUB_HSELEX"
    path = os.path.join(RHO, src, "Table S4.xlsx")
    rows = []
    book = wb(path)

    spec_map = {
        "List_BsRho_R14_Peaks": ("bsubtilis_168",     "168"),
        "List_EcRho_R10_Peaks": ("ecoli_K12_MG1655",  "MG1655"),
    }
    for sheet, (org, strain) in spec_map.items():
        for r in sheet_rows(book, sheet, skip=1):
            if not r[0]:
                continue
            peak_id = str(r[0]).strip()
            start   = int(r[1]) if r[1] is not None else None
            end     = int(r[2]) if r[2] is not None else None
            strand  = str(r[3]).strip() if r[3] else None
            seq     = str(r[4]).strip() if r[4] else None
            if start is None:
                continue
            rows.append(dict(
                source_id="BSUB_HSELEX", organism=org,
                strain=strain, site_id=peak_id,
                site_start=start, site_end=end, strand=strand,
                site_class="rut", terminates_or_silences="T",
                sequence=seq,
                evidence_type="T2_invitro_hselex",
                notes=f"sheet={sheet}",
            ))

    book.close()
    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


def parse_intrinsic_termite():
    """TERMITe 2024: genome-wide intrinsic terminator atlas — used as Rho decoys."""
    src  = "INTRINSIC_TERMITE"
    path = os.path.join(RHO, src, "Supplementary Table 2.xlsx")
    rows = []
    book = wb(path)

    # Columns: Species, chromosome, start, end, POT, sequence, strand, termite_id
    # Filter to E. coli and B. subtilis entries (they use "Chromosome" as chrom name)
    keep_species = {
        "Escherichia coli (a)": "ecoli_K12_MG1655",
        "Escherichia coli (b)": "ecoli_K12_MG1655",
        "Bacillus subtilis (a)": "bsubtilis_168",
        "Bacillus subtilis (b)": "bsubtilis_168",
        "Bacillus subtilis (c)": "bsubtilis_168",
        "Bacillus subtilis (d)": "bsubtilis_168",
    }

    for r in sheet_rows(book, "Atlas of intrinsic terminators", skip=2):
        if not r[0]:
            continue
        sp_raw = str(r[0]).strip()
        org    = keep_species.get(sp_raw)
        if org is None:
            continue
        strain = "MG1655" if "coli" in org else "168"
        start  = int(r[2]) if r[2] is not None else None
        end    = int(r[3]) if r[3] is not None else None
        pot    = int(r[4]) if r[4] is not None else None
        seq    = str(r[5]).strip() if r[5] else None
        strand = str(r[6]).strip() if r[6] else None
        tid    = str(r[7]).strip() if r[7] else None

        if start is None:
            continue
        rows.append(dict(
            source_id="INTRINSIC_TERMITE", organism=org,
            strain=strain, site_id=tid,
            site_start=start, site_end=end, strand=strand,
            site_class="intrinsic",
            terminates_or_silences="T",
            sequence=seq,
            evidence_type="T1_termseq",
            notes=f"POT={pot};species_label={sp_raw}",
        ))

    book.close()
    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


def parse_rhotermpredict():
    """RhoTermPredict: algorithm-derived RUT predictions for E. coli — cross-check only."""
    src  = "RHOTERMPREDICT"
    rows = []

    # MOESM1: full genome RUT predictions (Region, Start RUT, End RUT, Strand)
    path1 = os.path.join(RHO, src, "12859_2019_2704_MOESM1_ESM.xlsx")
    book1 = wb(path1)
    for r in sheet_rows(book1, "predictions", skip=1):
        if r[0] is None:
            continue
        region = str(r[0]).strip()
        start  = int(r[1]) if r[1] is not None else None
        end    = int(r[2]) if r[2] is not None else None
        strand = "+" if str(r[3]).lower() == "plus" else "-"
        if start is None:
            continue
        rows.append(dict(
            source_id="RHOTERMPREDICT", organism="ecoli_K12_MG1655",
            strain="MG1655", site_id=f"RUT_{region}_{strand}",
            site_start=start, site_end=end, strand=strand,
            site_class="rut_predicted",
            terminates_or_silences="T",
            sequence=None,
            evidence_type="algo_predicted_NEVER_HEADLINE",
            notes=f"source=MOESM1_full_genome",
        ))
    book1.close()

    # MOESM4: gene-associated predictions with read-count ratio
    path4 = os.path.join(RHO, src, "12859_2019_2704_MOESM4_ESM.xlsx")
    book4 = wb(path4)
    for r in sheet_rows(book4, "Foglio1", skip=2):
        if r[0] is None:
            continue
        start  = int(r[0]) if r[0] is not None else None
        end    = int(r[1]) if r[1] is not None else None
        ratio  = r[2]
        strand = "+" if str(r[3]).lower() == "plus" else "-"
        gene   = str(r[5]).strip() if r[5] else None
        if start is None:
            continue
        rows.append(dict(
            source_id="RHOTERMPREDICT", organism="ecoli_K12_MG1655",
            strain="MG1655", site_id=f"RUT_gene_{start}_{strand}",
            site_start=start, site_end=end, strand=strand,
            site_class="rut_predicted",
            terminates_or_silences="T",
            sequence=None,
            evidence_type="algo_predicted_NEVER_HEADLINE",
            notes=f"gene={gene};read_count_ratio={ratio};source=MOESM4_gene_assoc",
        ))
    book4.close()

    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# H. volcanii Ribosome Profiling — Gelsinger 2020 NAR 48:5201
# PMID 32382758, PMCID PMC7261190, DOI 10.1093/nar/gkaa210
# File: data/rbs_database/raw/HVOLC_RIBOSEQ/Ribo_MS_TableS1_final.xlsx
# Sheets: Annotated (1771), Unannotated TSS (18), smORF TSS (68),
#         Internal_Inframe (27), Internal_OutofFrame (16), N-terminal_extension (31)
# Annotated cols: Type, Start, Stop, Gene, Strand, Size(aa)
# Other sheets:   Type, Start, Stop, Strand, Size(aa), Start_codon, Peptide, Name, [Corresponding_gene]
# tis_pos = Start if + else Stop (1-based, H. volcanii DS2 GCF_000025685.1)
# ════════════════════════════════════════════════════════════════════════════

def parse_hvolc_riboseq():
    src = "HVOLC_RIBOSEQ"
    xlsx = os.path.join(RBS, src, "Ribo_MS_TableS1_final.xlsx")
    book = wb(xlsx)

    # --- Sheet: Annotated -------------------------------------------------
    # cols: Type(0), Start(1), Stop(2), Gene(3), Strand(4), Size_aa(5)
    ANNOTATED_SHEET = "Annotated"
    # --- Other sheets: same positional layout minus Gene col --------------
    OTHER_SHEETS = [
        "Unannotated TSS", "smORF TSS",
        "Internal_Inframe", "Internal_OutofFrame", "N-terminal_extension",
    ]

    rows = []

    def _add(type_label, start, stop, strand, locus_tag, gene_name, start_codon, notes_extra):
        try:
            start, stop = int(start), int(stop)
        except (TypeError, ValueError):
            return
        if strand not in ("+", "-"):
            return
        # Some sheets store minus-strand genes with Start > Stop (TIS at Start).
        # Use max/min so both orderings are handled correctly.
        tis_pos = min(start, stop) if strand == "+" else max(start, stop)
        tis_end = max(start, stop) if strand == "+" else min(start, stop)
        rows.append(dict(
            source_id=src,
            organism="hvolcanii_DS2",
            strain="DS2",
            gene_name=gene_name or "",
            locus_tag=locus_tag or "",
            tis_pos=tis_pos,
            tis_end=tis_end,
            strand=strand,
            start_codon=start_codon or "",
            ribo_density="",
            sd_motif="",
            sd_pos_rel="",
            evidence_type="T1",
            notes=f"type={type_label};{notes_extra}",
        ))

    # Annotated sheet
    for r in sheet_rows(book, ANNOTATED_SHEET, skip=1):
        type_label, start, stop, gene, strand, size_aa = r[0], r[1], r[2], r[3], r[4], r[5]
        _add(type_label, start, stop, strand,
             locus_tag=str(gene) if gene else "",
             gene_name=str(gene) if gene else "",
             start_codon="",
             notes_extra=f"size_aa={size_aa}")

    # Other sheets: Type(0), Start(1), Stop(2), Strand(3), Size_aa(4),
    #               Start_codon(5), Peptide(6), Name(7), [Corresponding_gene(8)]
    for sheet_name in OTHER_SHEETS:
        if sheet_name not in book.sheetnames:
            continue
        for r in sheet_rows(book, sheet_name, skip=1):
            type_label = r[0]
            start, stop, strand = r[1], r[2], r[3]
            size_aa  = r[4] if len(r) > 4 else None
            sc       = r[5] if len(r) > 5 else None
            name     = r[7] if len(r) > 7 else None
            corr     = r[8] if len(r) > 8 else None
            notes_extra = f"size_aa={size_aa}"
            if corr:
                notes_extra += f";corresponding_gene={corr}"
            _add(type_label, start, stop, strand,
                 locus_tag=str(name) if name else "",
                 gene_name=str(name) if name else "",
                 start_codon=str(sc) if sc else "",
                 notes_extra=notes_extra)

    book.close()
    out = os.path.join(RBS, src, f"{src.lower()}_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# E. coli Ribo-RET — Meydan et al. 2019 Mol Cell (PMID 30904393, PMC7115971)
# Files: data/rbs_database/raw/ECOLI_RIBORET/
#   TableS1_pTIS.xlsx          sheets: BW25113-pTISs, BL21-pTISs, Common iTISs,
#                                      BW25113-specific iTISs, BL21-specific iTISs
#   TableS2_Nterm_extensions.xlsx sheets: Common, BW25113-specific, BL21-specific
# SOUNDNESS FIX (2026-06-26): the original ad-hoc parse mixed in 3,608 BL21 rows.
#   BL21 is an E. coli B strain — its coordinates do NOT map to our K-12 genome
#   NC_000913.3, and no BL21 genome is in the panel. We therefore KEEP ONLY the
#   BW25113 (K-12 lineage) coordinates. BW25113 still differs from MG1655 by a
#   handful of indels → window-builder must liftover by gene name (noted per row).
# "Start position in BW25113" is already strand-aware (= start-codon position):
#   + strand start<stop, - strand start>stop. So tis_pos = Start directly.
# ════════════════════════════════════════════════════════════════════════════

def parse_ecoli_riboret():
    src = "ECOLI_RIBORET"
    d = os.path.join(RBS, src)
    rows = []

    def _add(tis_type, gene, start_codon, start, stop, strand,
             peak=None, rel_dens=None):
        if start in ("", None) or stop in ("", None) or strand not in ("+", "-"):
            return
        try:
            start, stop = int(start), int(stop)
        except (TypeError, ValueError):
            return
        note = f"type={tis_type};strain_coords=BW25113;liftover_to_MG1655_required"
        if peak not in (None, ""):
            note += f";peak_height={peak}"
        if rel_dens not in (None, ""):
            note += f";rel_density={rel_dens}"
        rows.append(dict(
            source_id=src, organism="ecoli_K12_MG1655", strain="BW25113",
            gene_name=str(gene) if gene else "", locus_tag="",
            tis_pos=start, tis_end=stop, strand=strand,
            start_codon=str(start_codon) if start_codon else "",
            ribo_density=rel_dens if rel_dens not in (None, "") else "",
            sd_motif="", sd_pos_rel="", evidence_type="T1_riboseq", notes=note,
        ))

    # ── TableS1: primary TIS + internal TIS (BW25113 only) ────────────────
    b1 = wb(os.path.join(d, "TableS1_pTIS.xlsx"))

    # BW25113-pTISs: Type, Gene, StartCodon, AAlen, Peak, Start, Stop, Strand
    for r in sheet_rows(b1, "BW25113-pTISs", skip=1):
        _add(r[0], r[1], r[2], r[5], r[6], r[7], peak=r[4])

    # BW25113-specific iTISs: Type,Gene,Upstream,StartCodon,AAlen,Peak,
    #                         Start,Stop,Strand,RelDensity
    for r in sheet_rows(b1, "BW25113-specific iTISs", skip=1):
        _add(r[0], r[1], r[3], r[6], r[7], r[8], peak=r[5], rel_dens=r[9])

    # Common iTISs: ...StartBW25113(8),StopBW25113(9),...,Strand(12),RelDensBW(13)
    for r in sheet_rows(b1, "Common iTISs", skip=1):
        _add(r[0], r[1], r[4], r[8], r[9], r[12], peak=r[6], rel_dens=r[13])
    b1.close()

    # ── TableS2: N-terminal extensions (BW25113 only) ─────────────────────
    b2 = wb(os.path.join(d, "TableS2_Nterm_extensions.xlsx"))
    # Common: Type,Gene,StartCodon,AAlen,PeakBW(4),PeakBL(5),StartBW(6),StopBW(7),
    #         StartBL(8),StopBL(9),Strand(10)
    for r in sheet_rows(b2, "Common", skip=1):
        _add(r[0], r[1], r[2], r[6], r[7], r[10], peak=r[4])
    # BW25113-specific: Type,Gene,StartCodon,AAlen,Peak,Start,Stop,Strand
    for r in sheet_rows(b2, "BW25113-specific", skip=1):
        _add(r[0], r[1], r[2], r[5], r[6], r[7], peak=r[4])
    b2.close()

    out = os.path.join(d, "ecoli_riboret_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# E. coli BL21(DE3) Ribo-RET — Meydan et al. 2019, BL21 arm of the same study
# Source xlsx live in ECOLI_RIBORET/ (TableS1_pTIS, TableS2_Nterm_extensions).
# Meydan mapped BL21 reads to CP001509.3 = NC_012971.2 (genome_build, GSE122129).
# We add BL21(DE3) as its own organism (genome dir ecoli_BL21_DE3) so coordinates
# anchor natively — validated: 2,753/2,753 pTIS land on exact start codons, so
# NO liftover is needed (unlike the BW25113 arm). A different E. coli lineage =
# extra TIS volume + a same-species cross-strain check. NOTE for splitting: BL21
# core genes are near-identical to K-12, so BL21 windows MUST be clustered with
# K-12 (MMseqs2/CD-HIT) to avoid train/val leakage (research_plan §5b).
# ════════════════════════════════════════════════════════════════════════════

def parse_ecoli_bl21_riboret():
    src = "ECOLI_BL21_RIBORET"
    d = os.path.join(RBS, "ECOLI_RIBORET")       # source xlsx shared with K-12 arm
    rows = []

    def _add(tis_type, gene, start_codon, start, stop, strand,
             peak=None, rel_dens=None):
        if start in ("", None) or stop in ("", None) or strand not in ("+", "-"):
            return
        try:
            start, stop = int(start), int(stop)
        except (TypeError, ValueError):
            return
        note = f"type={tis_type};strain_coords=BL21_DE3_native"   # no liftover
        if peak not in (None, ""):
            note += f";peak_height={peak}"
        if rel_dens not in (None, ""):
            note += f";rel_density={rel_dens}"
        rows.append(dict(
            source_id=src, organism="ecoli_BL21_DE3", strain="BL21_DE3",
            gene_name=str(gene) if gene else "", locus_tag="",
            tis_pos=start, tis_end=stop, strand=strand,
            start_codon=str(start_codon) if start_codon else "",
            ribo_density=rel_dens if rel_dens not in (None, "") else "",
            sd_motif="", sd_pos_rel="", evidence_type="T1_riboseq", notes=note,
        ))

    # ── TableS1: primary TIS + internal TIS (BL21 columns) ────────────────
    b1 = wb(os.path.join(d, "TableS1_pTIS.xlsx"))
    # BL21-pTISs: Type, Gene, StartCodon, AAlen, Peak, Start, Stop, Strand
    for r in sheet_rows(b1, "BL21-pTISs", skip=1):
        _add(r[0], r[1], r[2], r[5], r[6], r[7], peak=r[4])
    # BL21-specific iTISs: Type,Gene,Upstream,StartCodon,AAlen,Peak,Start,Stop,Strand,RelDensity
    for r in sheet_rows(b1, "BL21-specific iTISs", skip=1):
        _add(r[0], r[1], r[3], r[6], r[7], r[8], peak=r[5], rel_dens=r[9])
    # Common iTISs: ...StartBL21(10),StopBL21(11),Strand(12),PeakBL21(7),RelDensBL21(14)
    for r in sheet_rows(b1, "Common iTISs", skip=1):
        _add(r[0], r[1], r[4], r[10], r[11], r[12], peak=r[7], rel_dens=r[14])
    b1.close()

    # ── TableS2: N-terminal extensions (BL21 columns) ─────────────────────
    b2 = wb(os.path.join(d, "TableS2_Nterm_extensions.xlsx"))
    # Common: ...StartBL21(8),StopBL21(9),Strand(10),PeakBL21(5)
    for r in sheet_rows(b2, "Common", skip=1):
        _add(r[0], r[1], r[2], r[8], r[9], r[10], peak=r[5])
    # BL21-specific: Type,Gene,StartCodon,AAlen,Peak,Start,Stop,Strand
    for r in sheet_rows(b2, "BL21-specific", skip=1):
        _add(r[0], r[1], r[2], r[5], r[6], r[7], peak=r[4])
    b2.close()

    out = os.path.join(RBS, src, "ecoli_bl21_riboret_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# B. subtilis Ribo-seq — Lalanne et al. 2018 Cell (PMID 29606352, PMC5978003)
# File: data/rbs_database/raw/BSUB_RIBOSEQ/TableS1_synthesis_rates.xlsx
# Sheet "2_ B. subtilis": Gene, Start, Stop, Strand(1=fwd/0=rev),
#   Normalized ribosome footprint density, 25th pct, 75th pct  (header at row 3)
# Per the sheet legend: positions are forward-strand coords ordered low→high;
#   for the reverse strand Start/Stop refer to stop/start of the gene.
#   So tis_pos = Start if + else Stop. Density may be NaN or "[bracketed]"
#   (≤128 internal reads → low_reads flag).
# ════════════════════════════════════════════════════════════════════════════

def parse_bsub_riboseq():
    src = "BSUB_RIBOSEQ"
    d = os.path.join(RBS, src)
    book = wb(os.path.join(d, "TableS1_synthesis_rates.xlsx"))
    ws = book["2_ B. subtilis"]

    # Find header row (Gene/Start/Stop/Strand)
    header_idx = None
    grid = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(grid):
        if r and r[0] == "Gene" and r[1] == "Start" and r[2] == "Stop":
            header_idx = i
            break
    if header_idx is None:
        print(f"  [{src}] header not found in sheet '2_ B. subtilis'")
        book.close()
        return 0

    rows = []
    for r in grid[header_idx + 1:]:
        if not r or r[0] in (None, ""):
            continue
        gene, start, stop, strand_bit, dens = r[0], r[1], r[2], r[3], r[4]
        d25 = r[5] if len(r) > 5 else None
        d75 = r[6] if len(r) > 6 else None
        try:
            start, stop = int(start), int(stop)
        except (TypeError, ValueError):
            continue
        if strand_bit in (1, "1", 1.0):
            strand = "+"
        elif strand_bit in (0, "0", 0.0):
            strand = "-"
        else:
            continue
        tis_pos = start if strand == "+" else stop
        tis_end = stop  if strand == "+" else start

        # density: may be NaN or a "[bracketed]" low-read string
        low_reads = False
        dens_val = ""
        if dens not in (None, ""):
            s = str(dens).strip()
            if s.startswith("[") and s.endswith("]"):
                low_reads = True
                s = s[1:-1]
            if s.lower() not in ("nan", ""):
                try:
                    dens_val = float(s)
                except ValueError:
                    dens_val = ""

        rows.append(dict(
            source_id=src, organism="bsubtilis_168", strain="168",
            gene_name=str(gene), locus_tag="",
            tis_pos=tis_pos, tis_end=tis_end, strand=strand,
            start_codon="", ribo_density=dens_val,
            sd_motif="", sd_pos_rel="", evidence_type="T1_riboseq",
            notes=f"low_reads={low_reads};d25={d25};d75={d75}",
        ))
    book.close()

    out = os.path.join(d, "bsub_riboseq_tis_sites.tsv")
    return write_tsv(rows, RBS_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# E. coli BCM Rho-dependent terminators — Peters et al. 2012 Genes Dev
# PMID 23207917, PMCID PMC3521622, DOI 10.1101/gad.196741.112
# GEO GSE41936 (ChIP-chip) + GSE41939 (RNA-seq BCM)
# File: data/rho_database/raw/ECOLI_BCM_RHO/Supplemental_tables.xls
# Sheet "Table S1": Bicyclomycin Significant Transcripts (BSTs), 1264 rows
#   cols (row 17 header): start, end, length, # probes, strand, gene, type,
#     location, RNAseq (log2 +BCM/WT), rac_nusG, H-NS, MDS42_BCM, MDS42_nusG,
#     MDS42_nusA
# BSTs = regions where BCM (Rho inhibitor) increases RNA — i.e. Rho-terminated.
# Coordinates: E. coli K-12 MG1655 (NC_000913, 1-based).
# NOTE: research plan incorrectly called this "PNAS 109:15584" — correct journal
#   is Genes & Development. Contains both sense and antisense Rho terminators.
# ════════════════════════════════════════════════════════════════════════════

def parse_ecoli_bcm_rho():
    import xlrd
    src = "ECOLI_BCM_RHO"
    xls = os.path.join(RHO, src, "Supplemental_tables.xls")
    if not os.path.exists(xls):
        print(f"  [{src}] Supplemental_tables.xls not found — skipping")
        return 0

    book = xlrd.open_workbook(xls)
    ws = book.sheet_by_name("Table S1")

    # Find header row (contains 'start', 'end' in cols 0, 1)
    header_row = None
    for i in range(ws.nrows):
        if ws.cell_value(i, 0) == "start" and ws.cell_value(i, 1) == "end":
            header_row = i
            break
    if header_row is None:
        print(f"  [{src}] Could not find header row in Table S1")
        return 0

    rows = []
    for i in range(header_row + 1, ws.nrows):
        row = [ws.cell_value(i, j) for j in range(ws.ncols)]
        start = row[0]
        end   = row[1]
        length = row[2]
        strand = row[4]
        gene   = row[5]
        bst_type = row[6]   # sense / antisense / N/A
        location = row[7]   # end / within / UT
        rnaseq_fc = row[8]  # log2(+BCM/WT)

        if start in ("", None) or end in ("", None):
            continue
        try:
            start, end = int(start), int(end)
        except (TypeError, ValueError):
            continue
        if strand not in ("+", "-"):
            continue

        rows.append(dict(
            source_id=src,
            organism="ecoli_K12_MG1655",
            strain="MG1655",
            site_id=f"BST_{i - header_row:04d}",
            site_start=start,
            site_end=end,
            strand=strand,
            site_class=f"rho_BST_{bst_type}",
            terminates_or_silences="T",
            sequence=None,
            evidence_type="T1",
            notes=f"gene={gene};location={location};rnaseq_log2fc={rnaseq_fc:.3f};"
                  f"length_bp={length};method=BCM_RNAseq",
        ))

    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# MTB Rho-dependent TTS — Botella et al. 2022/2023 iScience
# PMID 37096044, PMCID PMC10122055, DOI 10.1016/j.isci.2023.106465
# ArrayExpress E-MTAB-11753 (raw BAMs only; processed sites are in SI)
# Files: data/rho_database/raw/MTB_RHODUC2/mmc2.xlsx – mmc8.xlsx
# mmc5.xlsx "High confidence RD TTS": 439 rows
#   cols: ID(0), Position(1), Strand(2), Locus(3), Gene name(4),
#         TTS type(5), Class(6), score(7)
#   TTS types: True_TTS (299), Cond_TTS (125), SecCond_TTS (15)
#   Class A = Rho-associated; O = other; I = intrinsic (not in this sheet)
# Position = exact bp 3'-end of nascent RNA (1-based, H37Rv NC_000962.3)
# Method: RhoDUC genetic Rho depletion (MTB Rho is BCM-resistant → orthogonal)
# NOTE: research plan cited "1,385 RD-TTS"; mmc5 has 439 high-confidence.
#   mmc4 "Classification of TTS" has 2568 total (class A=437 + P=184 + O=181 non-intrinsic).
# ════════════════════════════════════════════════════════════════════════════

def parse_mtb_rhoduc2():
    src = "MTB_RHODUC2"
    xlsx = os.path.join(RHO, src, "mmc5.xlsx")
    if not os.path.exists(xlsx):
        print(f"  [{src}] mmc5.xlsx not found — skipping")
        return 0

    book = wb(xlsx)
    ws = book["High confidence RD TTS"]
    # cols: ID(0), Position(1), Strand(2), Locus(3), Gene name(4),
    #        TTS type(5), Class(6), score(7)

    rows = []
    for r in sheet_rows(book, "High confidence RD TTS", skip=1):
        site_id  = r[0]
        pos      = r[1]
        strand   = r[2]
        locus    = r[3]
        gene     = r[4]
        tts_type = r[5]
        cls      = r[6]
        score    = r[7]

        try:
            pos = int(pos)
        except (TypeError, ValueError):
            continue
        if strand not in ("+", "-"):
            continue

        rows.append(dict(
            source_id=src,
            organism="mtuberculosis_H37Rv",
            strain="H37Rv",
            site_id=str(site_id),
            site_start=pos,
            site_end=pos,
            strand=strand,
            site_class=f"rho_TTS_{tts_type}",
            terminates_or_silences="T",
            sequence=None,
            evidence_type="T1",
            notes=f"locus={locus};gene={gene};class={cls};score={score};"
                  f"method=RhoDUC_genetic_depletion",
        ))

    book.close()
    out = os.path.join(RHO, src, f"{src.lower()}_sites.tsv")
    return write_tsv(rows, RHO_FIELDS, out)


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    print("=== RBS sources ===")
    totals = {}
    totals["ECOLI_RIBORET"] = parse_ecoli_riboret()
    totals["ECOLI_BL21_RIBORET"] = parse_ecoli_bl21_riboret()
    totals["BSUB_RIBOSEQ"]  = parse_bsub_riboseq()
    totals["ECOLI_TETRP"]   = parse_ecoli_tetrp()
    totals["CAULO_RIBOSEQ"] = parse_caulo_riboseq()
    totals["SAUR_EXSD"]     = parse_saur_exsd()
    totals["MTB_RIBOSEQ"]   = parse_mtb_riboseq()
    totals["BSUB_SPORE"]    = parse_bsub_spore()
    totals["HVOLC_RIBOSEQ"] = parse_hvolc_riboseq()

    print("\n=== Rho sources ===")
    totals["ECOLI_BCM_RHO"]     = parse_ecoli_bcm_rho()
    totals["MTB_RHODUC"]        = parse_mtb_rhoduc()
    totals["MTB_RHODUC2"]       = parse_mtb_rhoduc2()
    totals["BSUB_HSELEX"]       = parse_bsub_hselex()
    totals["INTRINSIC_TERMITE"] = parse_intrinsic_termite()
    totals["RHOTERMPREDICT"]    = parse_rhotermpredict()

    print("\n=== Summary ===")
    for src, n in totals.items():
        print(f"  {src:22s}: {n:>6,} rows")
    print(f"\nTotal: {sum(totals.values()):,} records across {len(totals)} sources")


if __name__ == "__main__":
    main()
